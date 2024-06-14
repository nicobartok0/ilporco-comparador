from openpyxl import load_workbook, Workbook
import cryptocode
import os


# Clase del scaner del libro de excel. El libro debe estar en el mismo
# directorio que el módulo.

class Lector:
    # Constructor de "Lector". Necesita el nombre del archivo excel del que sacar los datos.
    def __init__(self, name, ruta):
        self.name = name
        self.wb = load_workbook(ruta, data_only=True)
        self.ws = self.wb[self.wb.sheetnames[0]]
        self.intermedia = load_workbook('assets/tabla-intermedia.xlsx')
        self.wi = self.intermedia.worksheets[0]
        self.sku_list = []
        self.name_list = []
        self.cost_list = []
        self.code_list = []
        self.datos = {}
    
   

    # Método que toma los SKU's de la primer fila del excel.
    def obtener_skus(self):
        for column_data in self.ws['A']:
            if column_data.value != None:
                valor = str(column_data.value)
                self.sku_list.append(valor)
        return self.sku_list
    
    # Método que toma los nombres de los artículos de la tercer columna del excel.
    def obtener_nombres(self):
        for column_data in self.ws['C']:
            if column_data.value != None:
                valor = str(column_data.value)
                self.name_list.append(valor)
        return self.name_list
    
    # Método que toma los códigos de los artículos de la segunda columna del excel.
    def obtener_codigo(self):
        for column_data in self.ws['B']:
            if column_data.value != None and type(column_data.value) != str:
                valor = int(column_data.value)
                self.code_list.append(valor)
        return self.code_list
    
    def obtener_coste(self):
        for column_data in self.ws['E']:
            if column_data.value != None:
                valor = str(column_data.value)
                self.cost_list.append(valor)
        return self.cost_list
    


    # Método que toma los datos utilizando los métodos anteriores
    def obtener_datos(self):
        articulos = []
        self.obtener_skus()
        self.obtener_codigo()
        self.obtener_nombres()
        self.obtener_coste()
        for code in range(len(self.code_list)):
            articulo = {
                'nombre': f'{self.name_list[code]}',
                'SKU': f'{self.sku_list[code]}',
                'codigo': f'{self.code_list[code]}',
                'costo': f'{self.cost_list[code]}',
            }
            articulos.append(articulo)
        
        return articulos

    def intercode(self, articulos:dict):
        art_pricely = {}
        for row in self.wi.rows:
            if type(row[1].value) is int:
                art_pricely[row[1].value] = {
                    'CAGNOLI': row[3].value,
                    'NAHUEL': row[4].value,
                    'DOINA': row[5].value,
                    'LAS DINAS': row[6].value,
                    'CABAÑAS ARGENTINAS': row[7].value,
                    'OTROS': row[8].value
                }

        for articulo in articulos.values():
            if articulo.codigo in art_pricely.keys():
                for marca in articulo.cods_pricely.keys():
                    articulo.cods_pricely[marca] = art_pricely[articulo.codigo][marca]
            print(articulo.cods_pricely)
        return articulos


    def abrir_planilla_resultado(self):
        os.system(f'start excel.exe "{os.getcwd()}\\archivos\\{self.nombre_archivo}"')

